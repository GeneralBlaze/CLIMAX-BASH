import json
import os
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class TestGetTAGS():
    
    file_path = '/Users/princewill/Desktop/aTDO PIL/'
    excel_file_path = '/Users/princewill/Downloads/DON-CLIMAX/SHIPPING RECONCILLIATION/September docs/PIL/pil  Rough copy 3.xlsx'
    
    def setup_method(self, method):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1200')
        self.driver = webdriver.Chrome(options=options)
        self.driver.set_window_size(1157, 778)
        self.vars = {}

    def teardown_method(self, method):
        self.driver.quit()

    def log_not_found_bl(self, bl_number):
        not_found_file = '/Users/princewill/alx-interview/CLIMAX BASH/not_found1.json'
        if os.path.exists(not_found_file):
            with open(not_found_file, 'r') as file:
                not_found_data = json.load(file)
        else:
            not_found_data = []

        if bl_number not in not_found_data:
            not_found_data.append(bl_number)
            with open(not_found_file, 'w') as file:
                json.dump(not_found_data, file, indent=4)
        print(f"Logged not found BL number: {bl_number}")

    def log_failed_submission(self, bl_number):
        failed_submission_file = '/Users/princewill/alx-interview/CLIMAX BASH/failed_submission1.json'
        if os.path.exists(failed_submission_file):
            with open(failed_submission_file, 'r') as file:
                failed_submission_data = json.load(file)
        else:
            failed_submission_data = []

        if bl_number not in failed_submission_data:
            failed_submission_data.append(bl_number)
            with open(failed_submission_file, 'w') as file:
                json.dump(failed_submission_data, file, indent=4)
        print(f"Logged failed submission for BL number: {bl_number}")

    def update_excel(self, bl_number):
        wb = load_workbook(self.excel_file_path)
        sheet = wb.active
        amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):  # Assuming E column is the 5th column
            for cell in row:
                if cell.value == bl_number:
                    cell.fill = amber_fill
                    print(f"Updated Excel for BL number: {bl_number}")
                    break

        wb.save(self.excel_file_path)

    def test_getTAGS(self):
        """
        This test reads BL numbers and consignees from a JSON file, fills out a form on a webpage,
        uploads files, and submits the form. The form data, dropdown data, and consignee-specific
        data are also read from a JSON file to protect sensitive information.
        """
        # Read BL numbers and consignees from JSON
        with open('/Users/princewill/alx-interview/CLIMAX BASH/bl_data_TEST.json', mode='r') as file:
            bl_data = json.load(file)[0]
        
        # Read form data and dropdown data from JSON
        with open('/Users/princewill/alx-interview/CLIMAX BASH/form_data_TEST.json', mode='r') as file:
            form_config = json.load(file)
            form_data_template = form_config["form_data"]
            dropdown_data = form_config["dropdown_data"]
            consignee_data = form_config["consignee_data"]
            
            for consignee, bl_numbers in bl_data.items():
                for bl_number in bl_numbers:
                    file1 = os.path.join(self.file_path, f"{bl_number}.pdf")

                    # Check if the file exists
                    if not os.path.exists(file1):
                        self.log_not_found_bl(bl_number)
                        continue  # Skip to the next BL number

                    # Open the form page
                    self.driver.get("https://ecommerce.pilnigeria.com/refund-request")
                    print(f"Opened form page for BL number: {bl_number}")
                    
                    # Update form data with dynamic BL number and consignee-specific data
                    form_data = form_data_template.copy()
                    form_data.update(consignee_data[consignee])
                    form_data["edit-submitted-consignee-bank-details-bl-no-con"] = bl_number
                    form_data["edit-submitted-consignee-bank-details-upload-relevant-documents-fieldset-upload-document-con-description-of-document-con"] = f"Terminal Delivery Order for {bl_number}"

                    # Fill in the form fields using the dictionary
                    for field_id, value in form_data.items():
                        element = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.ID, field_id))
                        )
                        if value is not None:
                            element.send_keys(value)
                            print(f"Filled field {field_id} with value: {value}")
                        else:
                            element.click()
                            print(f"Clicked field {field_id}")
                        # Add a random delay to mimic human behavior
                        time.sleep(random.uniform(0.5, 2.0))
                    
                    # Handle dropdowns using the dictionary
                    for field_id, value in dropdown_data.items():
                        dropdown = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.ID, field_id))
                        )
                        dropdown.click()
                        option = dropdown.find_element(By.XPATH, f"//option[. = '{value}']")
                        option.click()
                        print(f"Selected dropdown {field_id} with value: {value}")
                        # Add a random delay to mimic human behavior
                        time.sleep(random.uniform(0.5, 2.0))

                    # Upload the files
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.ID, "edit-submitted-consignee-bank-details-upload-relevant-documents-fieldset-upload-document-con-upload-file-con-upload"))
                    ).send_keys(file1)
                    print(f"Uploaded file: {file1}")
                    # Add a random delay to mimic human behavior
                    time.sleep(random.uniform(0.5, 2.0))

                    # Submit the form
                    self.driver.find_element(By.NAME, "op").click()
                    print(f"Submitted form for BL number: {bl_number}")

                    # Wait for the form to submit and check for the "Go back to the form" link
                    try:
                        WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.LINK_TEXT, "Go back to the form"))
                        )
                        print(f"Form submission successful for BL number: {bl_number}")
                        self.driver.find_element(By.LINK_TEXT, "Go back to the form").click()
                    except:
                        print(f"Form submission failed for BL number: {bl_number}")
                        self.log_failed_submission(bl_number)

                    # Update Excel sheet
                    self.update_excel(bl_number)

# To run the test
if __name__ == "__main__":
    test = TestGetTAGS()
    test.setup_method(None)
    test.test_getTAGS()
    test.teardown_method(None)
    
    # Keep the browser open indefinitely
    print("Form filled. Keeping the browser open indefinitely for inspection.")
    while True:
        time.sleep(1)