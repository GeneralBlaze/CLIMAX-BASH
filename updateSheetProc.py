import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def color_matching_rows(json_path, folder_path, excel_path):
    """
    Colors the text in rows green in an Excel sheet if a match is found between tracking numbers
    from a JSON file and file names in a specified folder.

    Parameters:
    - json_path (str): Path to the JSON file containing tracking numbers.
    - folder_path (str): Path to the folder containing files to be checked.
    - excel_path (str): Path to the Excel file to be updated.
    """
    # Load tracking numbers from JSON
    with open(json_path, 'r') as json_file:
        tracking_numbers = json.load(json_file)

    # List all files in the specified folder
    files_in_folder = os.listdir(folder_path)

    # Open the Excel file
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # Define the green fill and font
    purple_font = Font(color="800080")

    # Iterate through the rows in column B
    for row in sheet.iter_rows(min_col=2, max_col=2, min_row=1):
        cell = row[0]
        for tracking_number in tracking_numbers:
            # Check if the tracking number is in any file name
            if any(tracking_number in file_name for file_name in files_in_folder):
                # Check if the tracking number matches the cell value in column B
                if tracking_number == cell.value:
                    # Color the text in the entire row green
                    for cell_in_row in sheet[cell.row]:
                        cell_in_row.font = purple_font

    # Save the updated Excel file
    workbook.save(excel_path)

# Example usage
json_path = r'/Users/princewill/alx-interview/CLIMAX BASH/trn.json'
folder_path = r'/Users/princewill/Downloads/test/merge _4_orient'
excel_path = r'/Users/princewill/alx-interview/CLIMAX BASH/oisa.xlsx'
color_matching_rows(json_path, folder_path, excel_path)