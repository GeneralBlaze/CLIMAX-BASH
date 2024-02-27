
import openpyxl

def get_formulas(file_path):
    wb = openpyxl.load_workbook(file_path)
    formulas = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.formula:
                    formulas[cell.coordinate] = cell.formula

    return formulas

def main():
    file_path = input('/Users/princewill/Downloads/DON-CLIMAX/DONCLIMAX ONNE WACT REFUND RECON adjusted format.xlsx')
    formulas = get_formulas(file_path)
    print("Formulas in the workbook:")
    for cell, formula in formulas.items():
        print(f"Cell {cell}: {formula}")

if __name__ == "__main__":
    main()
