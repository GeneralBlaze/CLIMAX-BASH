"""
This script merges PDF files based on mappings from an Excel sheet. It processes PDF files in a specified folder,
matches them with entries in an Excel sheet, and merges PDFs that correspond to the same entry in the Excel sheet.

The script performs the following steps:
1. Lists all PDF files in a specified folder.
2. Reads an Excel sheet and extracts mappings between two columns (B and C).
3. Matches PDF files with the Excel data based on the name of the PDF file and the value in column C.
4. Merges PDF files that match the same value in column B of the Excel sheet.
5. Moves processed PDF files to a 'processed' subfolder within the original folder.

The script assumes that the first row of the Excel sheet contains headers and starts processing from the fourth row.
"""

import os
from openpyxl import load_workbook
from PyPDF2 import PdfMerger
import shutil

# File paths
folder_path = r'/Users/princewill/alx-interview/CLIMAX BASH/msc cover docs/interchange'
excel_path = r'/Users/princewill/alx-interview/CLIMAX BASH/aisa.xlsx'

# Step 1: List all PDF files in the folder
pdf_files = [file for file in os.listdir(folder_path) if file.endswith('.pdf')]

# Step 2: Read the Excel sheet using openpyxl
wb = load_workbook(filename=excel_path)
sheet = wb.active

column_b_c_mapping = []
# Assuming the first row is headers
for row in sheet.iter_rows(min_row=4, values_only=True):
    column_b_c_mapping.append({'ColumnB': row[1], 'ColumnC': row[2]})
# print(f"Excel mappings: {column_b_c_mapping}")

# Step 3: Match PDF files with Excel data and note column B values
# Dictionary to hold column B values as keys and list of matching PDFs as values
pdf_to_merge = {}

for pdf_file in pdf_files:
    print(f"Processing PDF: {pdf_file}")
    for row in column_b_c_mapping:
        # Remove the '.pdf' extension from pdf_file before comparing
        pdf_file_without_extension = pdf_file.strip().lower().replace('.pdf', '')
        column_c_value_lower = str(row['ColumnC']).strip().lower()
        print(
            f"Comparing {pdf_file_without_extension} with {column_c_value_lower}")
        if pdf_file_without_extension == column_c_value_lower:
            print(f"Matching PDF: {pdf_file} with ColumnB: {row['ColumnB']}")
            if row['ColumnB'] in pdf_to_merge:
                pdf_to_merge[row['ColumnB']].append(pdf_file)
            else:
                pdf_to_merge[row['ColumnB']] = [pdf_file]

# Step 4 & 5: Merge PDF files for each group and save
processed_folder_path = os.path.join(folder_path, "processed")
if not os.path.exists(processed_folder_path):
    os.makedirs(processed_folder_path)

for column_b_value, pdfs in pdf_to_merge.items():
    print(f"Merging PDFs for {column_b_value}: {pdfs}")
    merger = PdfMerger()
    for pdf in pdfs:
        pdf_path = os.path.join(folder_path, pdf)
        if os.path.isfile(pdf_path): 
            print(f"Appending {pdf_path}")
            merger.append(pdf_path)
            # Move processed file to the new folder
            processed_pdf_path = os.path.join(processed_folder_path, pdf)
            shutil.move(pdf_path, processed_pdf_path)
            print(f"Moved {pdf_path} to {processed_pdf_path}")
        else:
            print(f"Warning: File {pdf_path} does not exist and will be skipped.")
    output_path = os.path.join(folder_path, f"{column_b_value}.pdf")
    merger.write(output_path)
    merger.close()
    print(f"Merged PDF saved as {output_path}")

print("PDF merging completed successfully.")