import asyncio
import json
import random
import openpyxl
from pyppeteer import launch
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Helper function to open a new page and navigate to tracking page
accepted_cookies = False


async def open_tracking_page(browser, tracking_number):
    global accepted_cookies
    page = await browser.newPage()
    await page.setViewport({'width': 1168, 'height': 845})

    await page.goto("https://www.msc.com/en/track-a-shipment", {'waitUntil': "networkidle0"})

    # Wait for the cookies popup to appear and click the "Accept" button
    if not accepted_cookies:
        try:
            await page.waitForXPath('//button[normalize-space()="Accept All"]', {'timeout': 30000})
            buttons = await page.xpath('//button[normalize-space()="Accept All"]')
            await buttons[0].click()
            accepted_cookies = True
        except Exception:
            print("No cookies popup found")

    await page.waitForSelector("#trackingNumber")
    await page.type("#trackingNumber", tracking_number)

    await asyncio.sleep(2)

    await page.waitForSelector(".msc-search-autocomplete__field > .msc-cta-icon-simple > .msc-icon-search")
    await page.click(".msc-search-autocomplete__field > .msc-cta-icon-simple > .msc-icon-search")

    await page.waitForSelector(".msc-flow-tracking__details-subtitle")
    return page

# Function to get number of containers


async def get_number_of_containers(page):
    try:
        no_of_containers = await page.querySelectorEval(".msc-flow-tracking__details-subtitle > span:nth-child(2)", "(el) => parseInt(el.innerText)")
        return no_of_containers
    except Exception as error:
        print(error)
        return None

# Function to run script for one container


async def run_one_container_script(page, tracking_number):
    container = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    vessel = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    voyage = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    one_pull_out = await page.querySelectorEval('.msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell--two .data-value', '(el) => el.innerText')
    one_return = await page.querySelectorEval('.msc-flow-tracking__step--over > .msc-flow-tracking__cell--two .data-value', '(el) => el.innerText')
    oneSize = await page.querySelectorEval(".msc-flow-tracking__cell--two .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")
    
    print(f'the size of the container is: {oneSize}')

    return {'container': container,
            'vessel': vessel,
            'voyage': voyage,
            'size': oneSize,
            'pull_out': one_pull_out,
            'return': one_return,
            'tracking_number': tracking_number}


# Function to run script for two containers


async def run_two_container_script(page, tracking_number):
    first_container = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    first_vessel = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    first_voyage = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    first_pull_out = await page.querySelectorEval('div:nth-child(4) .msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    first_return = await page.querySelectorEval('div:nth-child(4) .msc-flow-tracking__port:nth-child(2) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    first_size = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")

    second_container = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    second_vessel = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    second_voyage = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    second_pull_out = await page.querySelectorEval('div:nth-child(5) .msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    second_return = await page.querySelectorEval('div:nth-child(5) .msc-flow-tracking__port:nth-child(2) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    second_size = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__cell:nth-child(2) .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")

    # Create dictionaries for each container
    first_container_data = {
        'container': first_container,
        'vessel': first_vessel,
        'size': first_size,
        'voyage': first_voyage,
        'pull_out': first_pull_out,
        'return': first_return,
        'tracking_number': tracking_number
    }

    second_container_data = {
        'container': second_container,
        'vessel': second_vessel,
        'size': second_size,
        'voyage': second_voyage,
        'pull_out': second_pull_out,
        'return': second_return,
        'tracking_number': tracking_number
    }

    return [
        first_container_data, second_container_data
    ]

# Function to run script for three containers


async def run_three_container_script(page, tracking_number):
    first_container = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    first_vessel = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    first_voyage = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    first_pull_out = await page.querySelectorEval('div:nth-child(4) .msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    first_return = await page.querySelectorEval('div:nth-child(4) .msc-flow-tracking__port:nth-child(2) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    first_size = await page.querySelectorEval("div:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")

    second_container = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    second_vessel = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    second_voyage = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    second_pull_out = await page.querySelectorEval('div:nth-child(5) .msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    second_return = await page.querySelectorEval('div:nth-child(5) .msc-flow-tracking__port:nth-child(2) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    second_size = await page.querySelectorEval("div:nth-child(5) .msc-flow-tracking__cell:nth-child(2) .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")

    third_container = await page.querySelectorEval("div:nth-child(6) .msc-flow-tracking__cell:nth-child(1) .data-value", "(el) => el.innerText")
    third_vessel = await page.querySelectorEval("div:nth-child(6) .msc-flow-tracking__port:nth-child(4) .data-value:nth-child(2) > span:nth-child(2)", "(el) => el.innerText")
    third_voyage = await page.querySelectorEval("div:nth-child(6) .msc-flow-tracking__port:nth-child(4) span:nth-child(3)", "(el) => el.innerText")
    third_pull_out = await page.querySelectorEval('div:nth-child(6) .msc-flow-tracking__port:nth-child(4) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    third_return = await page.querySelectorEval('div:nth-child(6) .msc-flow-tracking__port:nth-child(2) .msc-flow-tracking__cell:nth-child(2) .data-value:nth-child(2)', 'element => element.innerText')
    third_size = await page.querySelectorEval("div:nth-child(6) .msc-flow-tracking__cell:nth-child(2) .msc-flow-tracking__data > div > .data-value", "(el) => el.textContent.match(/\d+/) ? el.textContent.match(/\d+/)[0] : ''")

    return [
        {'container': first_container, 'vessel': first_vessel,
            'voyage': first_voyage, 'pull_out': first_pull_out, 'return': first_return, 'tracking_number': tracking_number, 'size': first_size},
        {'container': second_container,
            'vessel': second_vessel, 'voyage': second_voyage, 'pull_out': second_pull_out, 'return': second_return, 'tracking_number': tracking_number, 'size': second_size},
        {'container': third_container, 'vessel': third_vessel,
            'voyage': third_voyage, 'pull_out': third_pull_out, 'return': third_return, 'tracking_number': tracking_number, 'size': third_size}
    ]

# Function to update Excel sheet


def update_excel_sheet(results):
    workbook = load_workbook(filename='SOA1a.xlsx')
    worksheet = workbook.active
    print(f'length result begining of update excel sheet {len(results)}')

    row_offset = 0  # Keep track of newly inserted rows to adjust the row numbers accordingly

    for base_row in range(6, 400 + row_offset):
        row = base_row + row_offset  # Adjust row number based on the offset
        tracking_number_cell = worksheet['B' + str(row)]
        tracking_number = tracking_number_cell.value

        if len(results) == 1:
            if tracking_number and results[0]['tracking_number'] == tracking_number:

                # Fill the first container data in the original row
                # Assign the first result dictionary to 'result'
                result = results[0]
                worksheet['C' + str(row)] = result['container']
                worksheet['D' + str(row)] = result['size']
                worksheet['E' + str(row)] = result['pull_out']
                worksheet['F' + str(row)] = result['return']
                worksheet['G' + str(row)] = result['vessel']
                worksheet['H' + str(row)] = result['voyage']

        elif len(results) > 1:
            # Iterate through each dictionary in 'results'
            for index in range(len(results)):
                result = results[index]  # Assign each dictionary to 'result'
                if result['tracking_number'] == tracking_number:
                    if index == 0:
                        # Fill the first container data in the original row
                        worksheet['C' + str(row)] = result['container']
                        worksheet['D' + str(row)] = result['size']
                        worksheet['E' + str(row)] = result['pull_out']
                        worksheet['F' + str(row)] = result['return']
                        worksheet['G' + str(row)] = result['vessel']
                        worksheet['H' + str(row)] = result['voyage']

                    else:
                        # Handle additional containers by inserting new rows
                        row_offset += 1  # Adjust for the new row
                        # Insert a new row below the current one
                        worksheet.insert_rows(row + 1)
                        # Copy formatting from the current row to the new row
                        # Assuming you want to copy all columns
                        for col in range(1, worksheet.max_column + 1):
                            new_cell = worksheet.cell(row=row + 1, column=col)
                            prev_cell = worksheet.cell(row=row, column=col)
                            if prev_cell.has_style:  # Check if the previous cell has any custom style to copy
                                new_cell.font = copy(prev_cell.font)
                                new_cell.border = copy(prev_cell.border)
                                new_cell.fill = copy(prev_cell.fill)
                                new_cell.number_format = copy(
                                    prev_cell.number_format)
                                new_cell.protection = copy(
                                    prev_cell.protection)
                                new_cell.alignment = copy(prev_cell.alignment)
                        # Update the row variable to point to the new row
                        row = row + 1
                        # Write the additional container data to the new row
                        # Copy tracking number
                        worksheet['B' + str(row)] = tracking_number
                        worksheet['C' + str(row)] = result['container']
                        worksheet['D' + str(row)] = result['size']
                        worksheet['E' + str(row)] = result['pull_out']
                        worksheet['F' + str(row)] = result['return']
                        worksheet['G' + str(row)] = result['vessel']
                        worksheet['H' + str(row)] = result['voyage']
    workbook.save('SOA1a.xlsx')


# Function to check if a tracking number already has a vessel and voyage number in the SOA sheet
def is_already_processed(tracking_number):
    # Open the SOA sheet
    workbook = openpyxl.load_workbook('SOA1a.xlsx')
    sheet = workbook.active

    # Iterate over the rows in the sheet
    for row in sheet.iter_rows(min_row=6, max_row=216):  # Rows 6 to 216
        # Check if the tracking number in the row matches the given tracking number
        if row[1].value == tracking_number:  # Column B
            # Check if the vessel and voyage number columns are not empty
            if row[5].value and row[6].value:  # Columns F and G
                return True

    return False

def remove_processed_tracking_numbers(processed_numbers):
    try:
        # Load the existing tracking numbers from the file
        with open("trn.json", "r") as file:
            trn = json.load(file)
        
        # Assuming trn is a list of tracking numbers
        # Remove the processed tracking numbers
        trn = [number for number in trn if number not in processed_numbers]
        
        # Write the updated list back to the file
        with open("trn.json", "w") as file:
            json.dump(trn, file)
            
    except Exception as error:
        print("Error processing tracking numbers:", error)

# Main function to run the appropriate script for each tracking number
async def main():
    try:
        with open("trn.json", "r") as file:
            trn = json.load(file)
    except Exception as error:
        print("Error loading tracking numbers:", error)
        return

    trn = list(set(trn))

    browser = await launch(headless=False)
    processed_count = 0  # Initialize the counter

    for tracking_number in trn:
        results = []  # Initialize results for each tracking number

        print(f"Processing tracking number {tracking_number}")
        if is_already_processed(tracking_number):
            print(
                f"Tracking number {tracking_number} already processed, skipping...")
            continue

        await asyncio.sleep(random.randint(5, 15))
        page = await open_tracking_page(browser, tracking_number)

        try:
            no_of_containers = await get_number_of_containers(page)

            if no_of_containers == 1:
                result = await run_one_container_script(page, tracking_number)
                results.append(result)
            elif no_of_containers == 2:
                containers = await run_two_container_script(page, tracking_number)
                results.extend(containers)
            elif no_of_containers == 3:
                containers = await run_three_container_script(page, tracking_number)
                results.extend(containers)
            else:
                print(
                    f"Tracking number {tracking_number}: Number of containers not supported by the script.")
        except Exception as error:
            print(f"Tracking number {tracking_number}:", error)
        finally:
            await page.close()

        try:
            update_excel_sheet(results)
            print("Excel sheet updated for tracking number", tracking_number)
            processed_count += 1  # Increment the counter after each tracking number is processed
            print(f"Total tracking numbers processed: {processed_count}")

            # Check if the processed_count is a multiple of 10
            if processed_count % 10 == 0:
                print("Processed 10 tracking numbers, pausing for 60 seconds...")
                await asyncio.sleep(60)  # Pause for 60 seconds
                print("Resuming processing...")
        except Exception as error:
            print("Error updating Excel sheet:", error)

    print(f"Total tracking numbers processed: {processed_count}")
    await browser.close()

if __name__ == "__main__":
    asyncio.get_event_loop().run_until_complete(main())
