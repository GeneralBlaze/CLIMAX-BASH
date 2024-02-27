import openpyxl

def copy_remark_column(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Get the BWACTT sheet
    source_sheet = wb['BWACTT']
    
    # Find the column index of the REMARK column
    remark_column_index = None
    for col in range(1, source_sheet.max_column + 1):
        if source_sheet.cell(row=3, column=col).value == 'REMARK':
            remark_column_index = col
            break
    
    if remark_column_index is None:
        print("REMARK column not found in BWACTT sheet")
        return
    
    # Create a new sheet for the REMARK column
    dest_sheet = wb.create_sheet(title='RemarkColumn')
    
    # Copy the REMARK column to the new sheet
    for row in range(3, source_sheet.max_row + 1):
        remark_value = source_sheet.cell(row=row, column=remark_column_index).value
        dest_sheet.cell(row=row - 2, column=1, value=remark_value)
    
    # Filter out rows with "OWING" in the REMARK column
    filtered_rows = [row for row in dest_sheet.iter_rows(min_row=1, max_row=dest_sheet.max_row, min_col=1, max_col=1) if row[0].value != "OWING"]
    dest_sheet.delete_rows(idx for idx, _ in filtered_rows)
    
    # Save the workbook
    wb.save(filename)

# Example usage
copy_remark_column("example.xlsx")
