from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

# Load the workbook
wb = load_workbook('test_sheet.xlsx', data_only=True)

# Get the BWACTT sheet
bwactt_sheet = wb['BWACTT']

# Prompt for the month
month = input("Enter the month: ")

# Create fill colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
amber_fill = PatternFill(start_color="AA6E15", end_color="AA6E15", fill_type="solid")

# Create font styles
bold_font = Font(bold=True)

# Create alignment styles
center_alignment = Alignment(horizontal='center')
right_alignment = Alignment(horizontal='right')

# Create border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Create named style for Naira currency formatting
naira_style = NamedStyle(name="naira_style")
naira_style.number_format = '₦#,##0.00'

# Dictionary to store unique consignees and their data
consignee_data_dict = {}

# Get column indices for required columns
required_columns = ['BILL LADING NUMBER', 'CONTAINER NUMBER', 'SIZES', 'UNUSED DAYS', 'REFUND AMOUNT', 'CONSIGNEE']
col_indices = {cell.value: idx+1 for idx, cell in enumerate(bwactt_sheet[3]) if cell.value in required_columns}

# Check if all required columns are found
if len(col_indices) != len(required_columns):
    missing_columns = set(required_columns) - set(col_indices.keys())
    print(f"Missing columns: {missing_columns}")
    exit()

# Iterate through rows starting from row 4
for row in bwactt_sheet.iter_rows(min_row=4, min_col=1, max_col=bwactt_sheet.max_column):
    consignee = row[col_indices['CONSIGNEE'] - 1].value
    consignee_data = {col: row[idx - 1].value for col, idx in col_indices.items()}
    consignee_data_dict.setdefault(consignee, []).append(consignee_data)

# Create a new workbook
new_wb = Workbook()

# Register the named style with the workbook
new_wb.add_named_style(naira_style)

# Iterate through unique consignees and create a new sheet for each
for consignee, data in consignee_data_dict.items():
    # Create sanitized sheet title
    consignee_sheet_title = ''.join(char if char.isalnum() else '_' for char in consignee)
    consignee_sheet = new_wb.create_sheet(title=consignee_sheet_title)
    headers = list(data[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = consignee_sheet.cell(row=1, column=col, value=header)
        cell.fill = amber_fill

    # Write data with filter for UNUSED DAYS >= 1
    for entry in data:
        if entry['UNUSED DAYS'] >= 1:
            row = [entry[col] for col in headers]
            consignee_sheet.append(row)

# Remove the default sheet
new_wb.remove(new_wb.active)

# Initialize dictionary to store refund sums
refund_sums = {}

# Calculate refund sums for each consignee
for consignee, data in consignee_data_dict.items():
    refund_sum = sum(entry['REFUND AMOUNT'] for entry in data if entry['UNUSED DAYS'] >= 1)
    refund_sums[consignee] = refund_sum

# Create a new sheet for SUMMARY
summary_sheet = new_wb.create_sheet(title="SUMMARY", index=0)
summary_sheet.append(["CONSIGNOR", "REFUND AMOUNT"])
summary_sheet['A1'].font = bold_font
summary_sheet['B1'].font = bold_font
summary_sheet['A1'].fill = amber_fill
summary_sheet['B1'].fill = amber_fill

# Write refund sums to SUMMARY sheet
for idx, (consignee, refund_sum) in enumerate(refund_sums.items(), start=2):
    summary_sheet.cell(row=idx, column=1, value=consignee)
    summary_sheet.cell(row=idx, column=2, value=refund_sum)

# Calculate total refund amount
total_refund = sum(refund_sums.values())

# Write total refund to SUMMARY sheet
summary_sheet.cell(row=len(refund_sums) + 2, column=1, value="Total Refund")
summary_sheet.cell(row=len(refund_sums) + 2, column=2, value=total_refund)
summary_sheet['A' + str(len(refund_sums) + 2)].font = bold_font
summary_sheet['B' + str(len(refund_sums) + 2)].font = bold_font

# Format cells in SUMMARY sheet
for row in summary_sheet.iter_rows(min_row=2, max_row=len(refund_sums) + 1):
    for cell in row:
        cell.alignment = center_alignment

# Iterate over all sheets in the workbook
for sheet in new_wb.sheetnames:
    ws = new_wb[sheet]

    # Insert 'S/N' column at the beginning
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value='S/N').font = bold_font
    ws['A1'].fill = amber_fill
    ws['A1'].alignment = center_alignment

    # Fill 'S/N' column with sequential numbers
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        row[0].value = idx
        row[0].alignment = center_alignment

    # Apply formatting to headers and data cells
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = bold_font
            cell.fill = amber_fill
            cell.alignment = center_alignment

    # Set column widths and apply border style
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
        for cell in column:
            cell.border = thin_border

    # Insert title at the beginning of each sheet
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=f"DONCLIMAX {month} 2024 WACT RECONCILIATION").font = bold_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws['A1'].fill = green_fill
    ws['A1'].alignment = center_alignment

# Save the new workbook
new_filename = f"DONCLIMAX {month} 2024 WACT RECONCILIATION.xlsx"
new_wb.save(new_filename)
print
