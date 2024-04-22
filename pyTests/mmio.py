from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('/Users/princewill/Downloads/mm2.xlsx')

# Get the active sheet
sheet = wb.active

# Initialize variables to track row indices
row_a = 2  # Starting row for column A
row_f = 2  # Starting row for column F

# Iterate through rows until column A is empty
while sheet.cell(row=row_a, column=1).value:
    # Get value from column A
    value_a = sheet.cell(row=row_a, column=1).value

    # Initialize a variable to track if a match is found
    match_found = False

    # Iterate through rows in column E to look for a match
    row_e = 2  # Starting row for column E
    while sheet.cell(row=row_e, column=5).value:
        # Get value from column E
        value_e = sheet.cell(row=row_e, column=5).value

        # Check if values match
        if value_a == value_e:
            match_found = True
            break

        # Move to the next row in column E
        row_e += 1

    # If no match was found, write value from column A to column F
    if not match_found:
        sheet.cell(row=row_f, column=6).value = value_a
        row_f += 1  # Move to the next row in column F

    # Move to the next row in column A
    row_a += 1

# Save the workbook
wb.save('outputn.xlsx')