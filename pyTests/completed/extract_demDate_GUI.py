import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json, re, time
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

def find_closest_date_dysm(sheet, start_row):
    """
    Find the closest date in the 'DYSM' sheet.
    """
    closest_date = None
    closest_row = None
    for row in range(start_row, 12999, -1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            cell_value_str = str(cell_value)
            if cell_value_str.startswith("SHIPPING PAYMENT-"):
                closest_date = cell_value_str.split('-')[1].strip()
                closest_row = row
                break
    return closest_date, closest_row

def find_closest_date_TERMINAL(sheet, start_row):
    """
    Find the closest date in the 'TERMINAL' sheet.
    """
    closest_date = None
    closest_row = None
    pattern = r'TERMINAL PAYMENT-(\d+)(?:ST|ND|RD|TH)\s+(\w+)\s+(\d+)'

    for row in range(start_row, 1, -1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            cell_value_str = str(cell_value)
            match = re.search(pattern, cell_value_str)
            if match:
                day = match.group(1)
                month_match = re.match(r'TERMINAL PAYMENT-\d+(?:ST|ND|RD|TH)\s+(\w{3})', cell_value_str)
                if month_match:
                    month_abbr = month_match.group(1)
                    month_num = datetime.strptime(month_abbr, "%b").strftime("%m")
                    year = match.group(3)
                    closest_date = f"{day}/{month_num}/{year}"
                    closest_row = row
                    break
    return closest_date, closest_row

def process_sheet(sheet, bill_number, bill_data, find_closest_date_func, demurrage_column, bill_column, sheet_name):
    """
    Process a sheet and update the bill data dictionary.
    """
    for cell in sheet[bill_column][13000:]:
        if cell.value == bill_number:
            demurrage_value = sheet.cell(row=cell.row, column=demurrage_column).value
            closest_date, _ = find_closest_date_func(sheet, cell.row)

            if bill_number not in bill_data:
                bill_data[bill_number] = []
            closest_date_formatted = None
            if closest_date:
                closest_date_formatted = datetime.strptime(closest_date, "%d/%m/%Y").strftime("%d-%b-%Y")
            bill_data[bill_number].append({'demurrage': demurrage_value, 'sheet': sheet_name, 'date': closest_date_formatted})

class SheetDialog(simpledialog.Dialog):
    def __init__(self, parent, title=None, sheet_names=None):
        self.sheet_names = sheet_names
        self.selection = None
        super().__init__(parent, title=title)

    def body(self, master):
        self.listbox = tk.Listbox(master)
        self.listbox.pack(fill="both", expand=True)

        for sheet_name in self.sheet_names:
            self.listbox.insert("end", sheet_name)

        return self.listbox

    def apply(self):
        try:
            self.selection = self.listbox.get(self.listbox.curselection())
        except tk.TclError:
            pass

def create_excel_sheet(bill_data, sheet_name):
    """
    Create a new Excel sheet and populate it with the bill data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["BILL NUMBER", "SHEET", "DATE", "DEMURRAGE", "CORRECTED"]
    ws.append(headers)

    for bill_number, data in bill_data.items():
        for item in data:
            ws.append([bill_number, item['sheet'], item['date'], item['demurrage'], None])

    # Formatting
    amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = amber_fill
        cell.font = Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.border = thin_border
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [1, 2, 3]:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    wb.save(f"{sheet_name}.xlsx")

def main():
    """
    Main function to process the Excel workbook and export the data.
    """
    bill_data = {}

    # Create a root window
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Show a message to the user before opening the file dialog
    messagebox.showinfo("Select Excel File", "Please select the excel file with the daily shipping and terminal sheets in the next window.")

    # Open a file dialog for the user to select the input file
    input_file_name = filedialog.askopenfilename()

    start_time = time.time()
    wb = openpyxl.load_workbook(input_file_name)
    print(f"Time taken to load workbook: {time.time() - start_time} seconds")

    # Extract the sheet names from the workbook
    sheet_names = wb.sheetnames

    # Present the sheet names as options for the user to select
    dysm_sheet_dialog = SheetDialog(root, title="Select 'Daily Shipping' sheet", sheet_names=sheet_names)
    terminal_sheet_dialog = SheetDialog(root, title="Select 'Daily Terminal' sheet", sheet_names=sheet_names)

    dysm_sheet = wb[dysm_sheet_dialog.selection]
    TERMINAL_sheet = wb[terminal_sheet_dialog.selection]

    # Show a message to the user before opening the second file dialog
    messagebox.showinfo("Select Excel File", "Please select the Excel file containing Ejison bill numbers in the next window.")

    # Open a second file dialog for the user to select the workbook for the 'LATE' sheet
    late_file_name = filedialog.askopenfilename()

    late_time = time.time()
    late_wb = openpyxl.load_workbook(late_file_name)
    print(f"Time taken to load second workbook: {time.time() - late_time} seconds")

    # Extract the sheet names from the 'LATE' workbook
    late_sheet_names = late_wb.sheetnames

    # Present the 'LATE' sheet names as options for the user to select
    late_sheet_dialog = SheetDialog(root, title="Select 'Ejison BL num.' sheet", sheet_names=late_sheet_names)

    late_sheet = late_wb[late_sheet_dialog.selection]

    # Prompt the user for the name of the new sheet
    new_sheet_name = simpledialog.askstring("Input", "Enter the name of the new sheet:")

    for late_row in late_sheet.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True):
        bill_number = late_row[0]

        if bill_number:
            process_sheet(dysm_sheet, bill_number, bill_data, find_closest_date_dysm, 10, 'E', 'Shipping')
            process_sheet(TERMINAL_sheet, bill_number, bill_data, find_closest_date_TERMINAL, 8, 'D', 'Terminal')

    create_excel_sheet(bill_data, new_sheet_name)

    print(f"Data exported to '{new_sheet_name}.xlsx'")
    
    root.destroy()  # Close tkinter window

if __name__ == "__main__":
    main()