import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime

def find_closest_date_dysm(sheet, start_row):
    """
    Find the closest date in the 'DYSM' sheet.
    """
    closest_date = None
    closest_row = None
    for row in range(start_row, 12999, -1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            cell_value_str = str(cell_value)
            if cell_value_str.startswith("SHIPPING PAYMENT-"):
                closest_date = cell_value_str.split('-')[1].strip()
                closest_row = row
                break
    return closest_date, closest_row

def find_closest_date_TERMINAL(sheet, start_row):
    """
    Find the closest date in the 'TERMINAL' sheet.
    """
    closest_date = None
    closest_row = None
    pattern = r'TERMINAL PAYMENT-(\d+)(?:ST|ND|RD|TH)\s+(\w+)\s+(\d+)'

    for row in range(start_row, 1, -1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            cell_value_str = str(cell_value)
            match = re.search(pattern, cell_value_str)
            if match:
                day = match.group(1)
                month_match = re.match(r'TERMINAL PAYMENT-\d+(?:ST|ND|RD|TH)\s+(\w{3})', cell_value_str)
                if month_match:
                    month_abbr = month_match.group(1)
                    month_num = datetime.strptime(month_abbr, "%b").strftime("%m")
                    year = match.group(3)
                    closest_date = f"{day}/{month_num}/{year}"
                    closest_row = row
                    break
    return closest_date, closest_row

def process_sheet(sheet, bill_number, bill_data, find_closest_date_func, demurrage_column, bill_column):
    """
    Process a sheet and update the bill data dictionary.
    """
    for cell in sheet[bill_column][13000:]:
        if cell.value == bill_number:
            demurrage_value = sheet.cell(row=cell.row, column=demurrage_column).value
            if demurrage_value:  # Only include entries with demurrage values
                closest_date, _ = find_closest_date_func(sheet, cell.row)
                sheet_name = sheet.title
                sheet_name = 'Shipping' if sheet_name == 'DYSM' else 'Terminal'  # Replace 'DYSM' with 'Shipping' and 'TERMINAL' with 'Terminal'

                if bill_number not in bill_data:
                    bill_data[bill_number] = []
                closest_date_formatted = None
                if closest_date:
                    closest_date_formatted = datetime.strptime(closest_date, "%d/%m/%Y").strftime("%d-%b-%Y")
                bill_data[bill_number].append({'demurrage': demurrage_value, 'sheet': sheet_name, 'date': closest_date_formatted})

def create_excel_sheet(bill_data):
    """
    Create a new Excel sheet and populate it with the bill data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["BILL NUMBER", "SHEET", "DATE", "DEMURRAGE", "CORRECTED"]
    ws.append(headers)

    for bill_number, data in bill_data.items():
        for item in data:
            ws.append([bill_number, item['sheet'], item['date'], item['demurrage'], None])


    # Formatting
    amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = amber_fill
        cell.font = Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in [1, 2, 3]:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    wb.save("bill_data.xlsx")

def main():
    """
    Main function to process the Excel workbook and export the data.
    """
    bill_data = {}

    wb = openpyxl.load_workbook('11 AND 26.xlsx')

    late_sheet = wb['LATE']
    dysm_sheet = wb['DYSM']
    TERMINAL_sheet = wb['TERMINAL']

    for late_row in late_sheet.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True):
        bill_number = late_row[0]

        if bill_number:
            process_sheet(dysm_sheet, bill_number, bill_data, find_closest_date_dysm, 10, 'E')
            process_sheet(TERMINAL_sheet, bill_number, bill_data, find_closest_date_TERMINAL, 8, 'D')

    create_excel_sheet(bill_data)

    print("Data exported to 'bill_data.xlsx'")

if __name__ == "__main__":
    main()