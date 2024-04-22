import openpyxl

def update_bwactt(file_path):
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the relevant sheets
    bwactt_sheet = workbook["BWACTT"]
    bchart_sheet = workbook["BCHART2"]
    ejison_sheet = workbook["EJISON2"]

    # Get the last row numbers for each sheet
    last_row_bwactt = bwactt_sheet.max_row
    last_row_bchart = bchart_sheet.max_row
    last_row_ejison = ejison_sheet.max_row

    # Iterate through each row in BWACTT
    for i in range(4, last_row_bwactt + 1):
        # If the bill laden number is not empty, skip this row
        if bwactt_sheet.cell(row=i, column=2).value:
            continue

        container_id = bwactt_sheet.cell(row=i, column=3).value
        bill_laden_number = ""

        # Look for a match in BCHART
        for j in range(6, last_row_bchart + 1):
            if bchart_sheet.cell(row=j, column=4).value == container_id:
                bill_laden_number = bchart_sheet.cell(row=j, column=3).value
                break

        # If no match in BCHART, look for a match in EJISON
        if not bill_laden_number:
            for k in range(6, last_row_ejison + 1):
                if ejison_sheet.cell(row=k, column=4).value == container_id:
                    bill_laden_number = ejison_sheet.cell(row=k, column=3).value
                    break

        # Update BWACTT with the retrieved bill number
        bwactt_sheet.cell(row=i, column=2).value = bill_laden_number

    # Save changes to the same Excel file
    workbook.save(file_path)
    print("Update complete.")

# Provide the path to your Excel file
file_path = "/Users/princewill/Downloads/Invoices and Recipts DCBTL/MT2.xlsx"

# Call the function to update BWACTT
update_bwactt(file_path)