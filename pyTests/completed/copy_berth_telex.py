import openpyxl

# Open both workbooks
wb1 = openpyxl.load_workbook('/Users/princewill/Downloads/DON-CLIMAX/Demmurage on Ejisonn.xlsx')
wb2 = openpyxl.load_workbook('/Users/princewill/Downloads/DON-CLIMAX/Book1.xlsx')

# Get the sheet from the first workbook
sheet1 = wb1.active

# Iterate over rows 874 to 1037 in column C of the first workbook's sheet
for row in range(461, 865):
    cell_value = sheet1.cell(row=row, column=3).value

    # Iterate over all the sheets in the second workbook
    for sheet2 in wb2.worksheets:
        # Iterate over all the cells in column D of the second workbook's sheet
        for cell in sheet2['D']:
            # If the cell value matches the cell value in column C of the first workbook's sheet
            if cell.value == cell_value:
                # Copy the corresponding cell values from columns H and J of the second workbook's sheet to columns D and E of the first workbook's sheet, respectively
                sheet1.cell(row=row, column=4).value = sheet2.cell(row=cell.row, column=8).value
                sheet1.cell(row=row, column=5).value = sheet2.cell(row=cell.row, column=10).value
                break

# Save the changes in the first workbook
wb1.save('/Users/princewill/Downloads/DON-CLIMAX/Demmurage on Ejisonn.xlsx')