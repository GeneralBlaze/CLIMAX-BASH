import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime

def find_closest_date_dysm(sheet, start_row):
    closest_date = None
    closest_row = None
    for row in range(start_row, 12999, -1):  # Iterate from start_row to row 1
        cell_value = sheet.cell(row=row, column=1).value  # Assuming the date is in the first column
        if cell_value:
            cell_value_str = str(cell_value)
            if cell_value_str.startswith("SHIPPING PAYMENT-"):
                closest_date = cell_value_str.split('-')[1].strip()
                closest_row = row
                break
    return closest_date, closest_row


def find_closest_date_term(sheet, start_row):
    closest_date = None
    closest_row = None
    pattern = r'(\d+)(?:ST|ND|RD|TH)\s+(\w+)\s+(\d+)'  # Regex pattern for the date format

    for row in range(start_row, 1, -1):  # Iterate from start_row to row 1
        cell_value = sheet.cell(row=row, column=1).value  # Assuming the date is in the first column
        if cell_value:
            cell_value_str = str(cell_value)
            match = re.search(pattern, cell_value_str)
            if match:
                day = match.group(1)
                month_abbr = re.match(r'\d+(?:ST|ND|RD|TH)\s+(\w{3})', cell_value_str).group(1)
                month_num = datetime.strptime(month_abbr, "%b").strftime("%m")
                year = match.group(3)
                closest_date = f"{day}/{month_num}/{year}"
                closest_row = row
                break
    return closest_date, closest_row

# Initialize a dictionary to store bill numbers and associated data
bill_data = {}

# Load the Excel workbook
wb = openpyxl.load_workbook('LATE TELEX.xlsx')

# Access the relevant sheets
late_sheet = wb['LATE']
dysm_sheet = wb['DYSM']
term_sheet = wb['TERM']

# Iterate through rows in the 'LATE' sheet (starting from row 5)
for late_row in late_sheet.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True):
    bill_number = late_row[0]

    if bill_number:
        # Check for matching bill numbers in the 'DYSM' sheet
        for dysm_cell in dysm_sheet['E'][13000:]:
            if dysm_cell.value == bill_number:
                # Collect demurrage value and sheet name
                demurrage_value = dysm_sheet.cell(row=dysm_cell.row, column=10).value
                closest_date, _ = find_closest_date_dysm(dysm_sheet, dysm_cell.row)
                sheet_name = 'Shipping'  # Change sheet name

                # Add data to the dictionary
                if bill_number not in bill_data:
                    bill_data[bill_number] = []
                closest_date_formatted = None
                if closest_date:
                    closest_date_formatted = datetime.strptime(closest_date, "%d/%m/%Y").strftime("%d-%b-%Y")
                bill_data[bill_number].append({'demurrage': demurrage_value, 'sheet': sheet_name, 'date': closest_date_formatted})

        # Check for matching bill numbers in the 'TERM' sheet
        for term_cell in term_sheet['E'][13000:]:
            if term_cell.value == bill_number:
                # Collect demurrage value and sheet name
                demurrage_value = term_sheet.cell(row=term_cell.row, column=9).value
                closest_date, _ = find_closest_date_term(term_sheet, term_cell.row)
                sheet_name = 'Terminal'  # Change sheet name

                # Add data to the dictionary
                if bill_number not in bill_data:
                    bill_data[bill_number] = []
                closest_date_formatted = None
                if closest_date:
                    closest_date_formatted = datetime.strptime(closest_date, "%d/%m/%Y").strftime("%d-%b-%Y")
                bill_data[bill_number].append({'demurrage': demurrage_value, 'sheet': sheet_name, 'date': closest_date_formatted})

# Save bill data to a JSON file
with open('bill_data.json', 'w') as json_file:
    json.dump(bill_data, json_file, indent=4)

print("Data exported to 'bill_data.json'")

# Load the bill data from JSON file
with open('bill_data.json') as json_file:
    bill_data = json.load(json_file)

# Create a new workbook
wb_new = openpyxl.Workbook()

# Create a new sheet for the data
sheet_new = wb_new.active
sheet_new.title = 'New Sheet'

# Set the title row data with the rearranged columns
title_row = ['Bill Number', 'Narration', 'Date', 'Demurrage', 'Corrected']
sheet_new.append(title_row)

# Apply formatting to the title row
for cell in sheet_new[1]:
    cell.font = Font(size=14, bold=True)
    cell.fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")

# Iterate over the bill data and fill the new sheet with the rearranged columns
for bill_number, entries in bill_data.items():
    for entry in entries:
        narration = 'Shipping' if entry['sheet'] == 'DYSM' else 'Terminal' if entry['sheet'] == 'TERM' else entry['sheet']
        row_data = [bill_number, narration, entry['date'], entry['demurrage'], None]  # Add None for the "Corrected" column
        sheet_new.append(row_data)

# Iterate over each column and calculate the maximum width needed
for col in range(1, sheet_new.max_column + 1):
    max_length = 0
    column = get_column_letter(col)
    for row in range(1, sheet_new.max_row + 1):
        cell = sheet_new[column + str(row)]
        if cell.value:
            cell_length = len(str(cell.value)) + 2  # Add padding
            if cell_length > max_length:
                max_length = cell_length
    sheet_new.column_dimensions[column].width = max_length

# Save the modified workbook
wb_new.save('bill_data_sheet.xlsx')

print("Data populated to the new sheet.")
