from openpyxl import load_workbook
from openpyxl import Workbook
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Alignment, PatternFill, Alignment, Border, Side, NamedStyle
import os
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Load the workbook
wb = load_workbook('test_sheet.xlsx', data_only=True)

# Get the BWACTT sheet
bwactt_sheet = wb['BWACTT']

# Prompt for the month
month = input("Enter the month: ")

# Create a fill for the background color
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

 #Create a fill for the background color
amber_fill = PatternFill(start_color="AA6E15", end_color="AA6E15", fill_type="solid")

# Create a font for the bold text
bold_font = Font(bold=True)

# Create an alignment for the right-aligned text
right_alignment = Alignment(horizontal='right')
left_alignment = Alignment(horizontal='left')


# Create a border for all cells
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Create a named style for the Naira currency
naira_style = NamedStyle(name="naira_style")
naira_style.number_format = '₦#,##0.00'


# Dictionary to store unique consignees and their data
consignee_data_dict = {}

# Get the column indices for the required columns
columns = ['BILL LADING NUMBER', 'CONTAINER NUMBER', 'SIZES', 'UNUSED DAYS', 'REFUND AMOUNT', 'CONSIGNEE']
col_indices = {}
for col_idx, cell in enumerate(bwactt_sheet[3], start=1):
    if cell.value in columns:
        col_indices[cell.value] = col_idx

# Check if all required columns are found
if len(col_indices) != len(columns):
    
    missing_columns = set(columns) - set(col_indices.keys())
    print(f"Missing columns: {missing_columns}")
    exit()

# Iterate through rows starting from row 4
for row in bwactt_sheet.iter_rows(min_row=4, min_col=1, max_col=bwactt_sheet.max_column):
    
    # Check if CONSIGNEE column index is present
    if 'CONSIGNEE' not in col_indices:
        print("CONSIGNEE column not found.")
        continue

    consignee = row[col_indices['CONSIGNEE'] - 1].value
    consignee_data = {}
    for col, idx in col_indices.items():
        consignee_data[col] = row[idx - 1].value

    if consignee in consignee_data_dict:
        consignee_data_dict[consignee].append(consignee_data)
    else:
        consignee_data_dict[consignee] = [consignee_data]


# Create a JSON file with the consignee details
with open('consignee_details.json', 'w') as json_file:
    json.dump(consignee_data_dict, json_file)

# Create a new workbook
new_wb = Workbook()

# Register the named style with the workbook
new_wb.add_named_style(naira_style)

# A counter to ensure unique sanitized names
counter = 1

# Define the consignee_name_map dictionary outside the loop
consignee_name_map = {}

# Iterate through unique consignees and create a new sheet for each
for consignee, data in consignee_data_dict.items():
    # Replace special characters with whitespace
    consignee_sheet_title = consignee.replace('!', '').replace('@', '').replace('#', '').replace('$', '').replace('%', '').replace('^', '').replace('&', '').replace('*', '').replace('(', '').replace(')', '').replace('-', '').replace('_', '').replace('+', '').replace('=', '').replace('[', '').replace('{', '').replace(']', '').replace('}', '').replace('|', '').replace(':', '').replace(';', '').replace('"', '').replace("'", '').replace('<', '').replace('>', '').replace('.', '').replace('?', '').replace('/', '').replace('\\', '').replace('~', '').replace('`', '').replace(',', '')

    # If the sanitized name already exists in the map, append a unique identifier to it
    temp_title = consignee_sheet_title
    while temp_title in consignee_name_map:
        temp_title = consignee_sheet_title + str(counter)
        counter += 1
    consignee_sheet_title = temp_title

    # Add the sanitized consignee and original consignee to the consignee_name_map
    consignee_name_map[consignee_sheet_title] = consignee

    # Create a new sheet if it doesn't already exist
    if consignee_sheet_title not in new_wb.sheetnames:
        consignee_sheet = new_wb.create_sheet(title=consignee_sheet_title)
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            consignee_sheet.cell(row=1, column=col, value=header)
            # Set the color to amber
            consignee_sheet.cell(row=1, column=col).fill = amber_fill
    else:
        consignee_sheet = new_wb[consignee_sheet_title]

    # Write data with filter for UNUSED DAYS >= 1
    row_idx = len(consignee_sheet['A']) + 1
    for entry in data:
        try:
            refund_amount = float(entry['REFUND AMOUNT'])
            if entry['UNUSED DAYS'] >= 1:
                for col_idx, value in enumerate(entry.values(), start=1):
                    cell = consignee_sheet.cell(row=row_idx, column=col_idx, value=value)
                    # Adjust column widths based on content length with padding
                    col_letter = get_column_letter(col_idx)
                    col_width = max(len(str(value)) + 2, len(col_letter) + 2)  # Padding of 2
                    consignee_sheet.column_dimensions[col_letter].width = col_width
                row_idx += 1
        except ValueError:
            pass


# Remove the default sheet
new_wb.remove(new_wb.active)

# Initialize an empty dictionary to store refund sums
refund_sums = {}

for sheet_name in new_wb.sheetnames:
    sheet = new_wb[sheet_name]
    refund_sum = 0
    # Use the consignee_name_map to get the original consignee name
    original_consignee = consignee_name_map[sheet_name]
    data = consignee_data_dict.get(original_consignee, [])
    for entry in data:
        refund_amount = entry['REFUND AMOUNT']
        unused_days = entry['UNUSED DAYS']
        try:
            refund_amount = float(refund_amount)
            unused_days = int(unused_days)
            if unused_days >= 1 and refund_amount >= 0:
                refund_sum += refund_amount
        except ValueError:
            pass

    # Add the sheet name and refund sum to the dictionary
    refund_sums[sheet_name] = refund_sum

# Write the refund sums to a JSON file
with open('refund_sums.json', 'w') as f:
    json.dump(refund_sums, f)

# Now, read from the JSON file to fill the sum total cell
with open('refund_sums.json', 'r') as f:
    refund_sums = json.load(f)

for sheet_name, refund_sum in refund_sums.items():
    sheet = new_wb[sheet_name]

    # Find the current index of the 'REFUND AMOUNT' column
    for i, col in enumerate(sheet.iter_cols(values_only=True)):
        if col[0] == 'REFUND AMOUNT':
            refund_col_idx = i
            break

    last_row = sheet.max_row + 2

    # Add total refund amount
    total_refund_cell = sheet.cell(row=last_row, column=refund_col_idx + 1)
    total_refund_cell.value = float(refund_sum)

    # Set text color to red
    red_font = Font(color=Color(rgb='00FF0000'))
    total_refund_cell.font = red_font

# SUMMARY sheet
refund_sum = 0
# Use the consignee_name_map to get the original consignee name
original_consignee = consignee_name_map[sheet_name]
data = consignee_data_dict.get(original_consignee, [])
for entry in data:
    refund_amount = entry['REFUND AMOUNT']
    unused_days = entry['UNUSED DAYS']
    try:
        refund_amount = float(refund_amount)
        unused_days = int(unused_days)
        if unused_days >= 1 and refund_amount >= 0:
            refund_sum += refund_amount
    except ValueError:
        pass

# Add the sheet name and refund sum to the dictionary
refund_sums[sheet_name] = refund_sum

# Write the refund sums to a JSON file
with open('refund_sums.json', 'w') as f:
    json.dump(refund_sums, f)

# Now, read from the JSON file to fill the sum total cell
with open('refund_sums.json', 'r') as f:
    data = json.load(f)

# Create a new sheet at the beginning of the workbook
summary_sheet = new_wb.create_sheet("SUMMARY", 0)

# Add the headers
summary_sheet.append([ "CONSIGNEE", "REFUND AMOUNT"])

# Fill the sheet with data from the JSON file
for i, (consignee, refund_amount) in enumerate(data.items(), start=1):
    summary_sheet.append([ consignee, refund_amount])
   
# Background color for the headers and alignment
for cell in summary_sheet[1]:
        cell.fill = amber_fill
        cell.alignment = Alignment(horizontal='center')

# Center align cell values
# Iterate through each row starting from row 2
for row in summary_sheet.iter_rows(min_row=2):
    for cell in row:
        # Get the title from the corresponding cell in the first row
        title_cell = summary_sheet.cell(row=1, column=cell.column)
        title = title_cell.value
        
        # Align the cell based on the title
        if title == "REFUND AMOUNT":
            cell.alignment = Alignment(horizontal='right')
        elif title in ["S/N", "CONSIGNEE"]:
            cell.alignment = Alignment(horizontal='center')

        

# Calculate the sum total of the refunds
total_refund = sum(data.values())

# Write the total refund to the last cell
last_row = summary_sheet.max_row + 2
total_refund_cell = summary_sheet.cell(row=last_row, column=2)
total_refund_cell.value = float(total_refund)

# Adjust column widths
for column in summary_sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    summary_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


# Iterate over all sheets in the workbook
for sheet in new_wb.worksheets:

    # Insert a column at the beginning of the sheet for 'S/N'
    sheet.insert_cols(1)
    sheet['A1'].value = 'S/N'  # Add a title to the 'S/N' column
    sheet['A1'].font = bold_font  # Make the title bold
    sheet['A1'].fill = amber_fill  # Add a background color to the title
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        row[0].value = i  # Fill the 'S/N' column with sequential numbers
        row[0].alignment = Alignment(horizontal='center')  # Center align the text

    # Make the title row bold
    for cell in sheet[1]:
        cell.font = bold_font

    # Find the 'BILL LADING NUMBER' column, align it to the left and adjust its width
    def format_header(sheet, column_index, font, alignment):
        """Format the header cell of a given column."""
        header_cell = sheet.cell(row=1, column=column_index)
        header_cell.font = font
        header_cell.alignment = alignment

    def format_cells(sheet, column_index, style, alignment):
        """Format the cells under the header of a given column."""
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, values_only=False):
            for row_cell in cell[1:]:  # Skip the header cell
                if style is not None:
                    row_cell.style = style
                row_cell.alignment = alignment

    def adjust_column_width(sheet, column_index, width):
        """Adjust the width of a given column."""
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    # Iterate over columns and apply formatting
    for i, col in enumerate(sheet.iter_cols(values_only=True), 1):  # Start counting from 1
        if col[0] in ["UNUSED DAYS", "SIZES", "CONTAINER NUMBER", "BILL LADING NUMBER"]:
            format_header(sheet, i, bold_font, Alignment(horizontal='center'))
            format_cells(sheet, i, None, Alignment(horizontal='center'))
            adjust_column_width(sheet, i, 15)
        elif col[0] == 'REFUND AMOUNT':
            format_header(sheet, i, bold_font, Alignment(horizontal='center'))
            format_cells(sheet, i, naira_style, Alignment(horizontal='right'))

    # Add borders to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Insert a row at the beginning of the sheet
    sheet.insert_rows(1)

    # Write the text to the first cell
    cell = sheet.cell(row=1, column=1)
    cell.value = f"DONCLIMAX {month} 2024 WACT RECONCILIATION"
    
    #sheet name for saving
    s_name = cell.value
    new_name = f"{s_name}.xlsx"

    # Merge cells
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

    # Format the merged cell
    cell.fill = green_fill
    cell.font = bold_font

    # Center align the text
    cell.alignment = Alignment(horizontal='center')



# Prompt to delete the JSON files
delete_files = input("Do you want to delete the JSON files? (y/n): ")

if delete_files.lower() == 'y':
    try:
        os.remove('consignee_details.json')
        os.remove('refund_sums.json')
        print("Files deleted successfully.")
    except Exception as e:
        print(f"Error occurred while deleting files: {e}")

# Save the new workbook
new_wb.save(new_name)
print("Operation completed")